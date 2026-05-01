import os
import json
import math
import re
import threading
from pathlib import Path
from dataclasses import dataclass
from typing import List, Dict, Any, Optional

try:
    from groq import Groq
except ImportError:
    Groq = None


APP_CONFIG_PATH = Path.home() / ".simplex_groq_solver.json"


@dataclass
class Constraint:
    name: str
    coeffs: List[float]
    sense: str
    rhs: float


@dataclass
class LPModel:
    problem_name: str
    objective_sense: str
    variables: List[str]
    objective_coeffs: List[float]
    constraints: List[Constraint]
    nonnegativity: bool = True


def _as_float(x):
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", ".")
    return float(s)


def _extract_json_object(text: str) -> Dict[str, Any]:
    """Extrae un JSON aunque el modelo lo devuelva dentro de ```json ... ```."""
    if text is None:
        raise ValueError("La respuesta del modelo vino vacia")

    clean = text.strip()
    clean = re.sub(r"^```(?:json)?", "", clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r"```$", "", clean).strip()

    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        start = clean.find("{")
        end = clean.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(clean[start:end + 1])
        raise


def load_app_config() -> Dict[str, Any]:
    if not APP_CONFIG_PATH.exists():
        return {}

    try:
        with APP_CONFIG_PATH.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (OSError, json.JSONDecodeError):
        return {}

    return data if isinstance(data, dict) else {}


def save_app_config(data: Dict[str, Any]) -> None:
    with APP_CONFIG_PATH.open("w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=True, indent=2)


def get_groq_api_key() -> str:
    env_api_key = os.environ.get("GROQ_API_KEY", "").strip()
    if env_api_key:
        return env_api_key

    config_api_key = str(load_app_config().get("groq_api_key", "")).strip()
    if config_api_key:
        return config_api_key

    return ""


def set_groq_api_key(api_key: str) -> None:
    clean_api_key = api_key.strip()
    config = load_app_config()

    if clean_api_key:
        config["groq_api_key"] = clean_api_key
    else:
        config.pop("groq_api_key", None)

    save_app_config(config)


def extract_lp_with_groq(problem_text: str, model: Optional[str] = None) -> LPModel:
    """
    Usa Groq para convertir el enunciado en un modelo de programacion lineal.

    IMPORTANTE:
    - Usa response_format={"type": "json_object"}, que es mas compatible.
    - Si quieres usar json_schema, cambia el modelo a uno que Groq soporte para structured outputs.
    """
    if Groq is None:
        raise RuntimeError("Instala la libreria de Groq con: pip install groq")

    api_key = get_groq_api_key()
    if not api_key:
        raise RuntimeError(
            "Falta la API key de Groq. Configurala en la ventana de la aplicacion "
            "o define GROQ_API_KEY en el sistema."
        )

    client = Groq(api_key=api_key)

    system = """
Eres experto en Investigacion de Operaciones y metodo simplex.
Tu tarea es extraer un modelo de programacion lineal desde un enunciado en espanol.

Devuelve UNICAMENTE JSON valido, sin markdown, sin comentarios y sin explicaciones.

El JSON debe tener exactamente esta estructura:
{
  "problem_name": "nombre corto del problema",
  "objective_sense": "max" o "min",
  "variables": ["x1 = descripcion", "x2 = descripcion"],
  "objective_coeffs": [coeficiente_x1, coeficiente_x2],
  "constraints": [
    {
      "name": "nombre de la restriccion",
      "coeffs": [coeficiente_x1, coeficiente_x2],
      "sense": "<=", ">=" o "=",
      "rhs": valor_derecha
    }
  ]
}

Reglas importantes:
- Define variables de decision claras.
- Si el problema busca ingreso, utilidad, beneficio o ganancia maxima, usa objective_sense = "max".
- Si el problema da precio de venta y costo de fabricacion, la utilidad por unidad es:
  utilidad = precio_de_venta - costo_de_fabricacion.
  Ejemplo: precio 52 y costo 30 => coeficiente objetivo 22.
  Ejemplo: precio 48 y costo 28 => coeficiente objetivo 20.
- En objective_coeffs coloca UN SOLO coeficiente por variable.
  Si hay 2 variables, objective_coeffs debe tener exactamente 2 numeros.
  Nunca coloques listas como [precio, costo] ni expresiones separadas.
- Si busca costo minimo, usa objective_sense = "min".
- La longitud de objective_coeffs debe coincidir con la cantidad de variables.
- La longitud de coeffs en cada restriccion debe coincidir con la cantidad de variables.
- No incluyas restricciones de no negatividad dentro de constraints; se asume xi >= 0.
- Convierte frases de demanda a restricciones lineales.
- Ejemplo: "interiores no puede ser mayor que exteriores en mas de 1" significa: interior - exterior <= 1.
  Si x1 = exterior y x2 = interior, entonces: -x1 + x2 <= 1, coeffs = [-1, 1].
""".strip()

    chosen_model = model or os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")

    completion = client.chat.completions.create(
        model=chosen_model,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": problem_text},
        ],
        temperature=0,
        response_format={"type": "json_object"},
    )

    content = completion.choices[0].message.content
    data = _extract_json_object(content)
    return lp_from_dict(data)


def _repair_objective_coeffs_if_needed(raw_coeffs: List[Any], variables: List[str]) -> List[float]:
    """
    Repara errores comunes de Groq cuando devuelve mas coeficientes que variables.

    Caso comun:
    - El problema da precio de venta y costo.
    - Groq devuelve [52, 0, 30, 48, 0, 28] para dos variables.
    - Eso realmente significa (52 - 30) y (48 - 28), por tanto [22, 20].

    Esta funcion no reemplaza la validacion; solo corrige patrones evidentes.
    """
    coeffs = [_as_float(v) for v in raw_coeffs]
    n = len(variables)

    if len(coeffs) == n:
        return coeffs

    # Patron: [precio1, 0, costo1, precio2, 0, costo2, ...]
    # Viene de interpretar "precio - costo" como tres elementos.
    if len(coeffs) == 3 * n:
        repaired = []
        for i in range(n):
            precio = coeffs[3 * i]
            costo = coeffs[3 * i + 2]
            repaired.append(precio - costo)
        print("\nAVISO: Groq devolvio objective_coeffs con longitud incorrecta.")
        print(f"Se interpreto como precio - costo y se corrigio a: {repaired}")
        return repaired

    # Patron: [precio1, costo1, precio2, costo2, ...]
    if len(coeffs) == 2 * n:
        repaired = []
        for i in range(n):
            precio = coeffs[2 * i]
            costo = coeffs[2 * i + 1]
            repaired.append(precio - costo)
        print("\nAVISO: Groq devolvio objective_coeffs con longitud incorrecta.")
        print(f"Se interpreto como precio - costo y se corrigio a: {repaired}")
        return repaired

    raise ValueError(
        "objective_coeffs debe tener la misma longitud que variables. "
        f"Variables: {len(variables)}. Coeficientes: {len(coeffs)}. "
        f"Coeficientes recibidos: {coeffs}"
    )


def lp_from_dict(data: Dict[str, Any]) -> LPModel:
    variables = [str(v) for v in data["variables"]]
    c = _repair_objective_coeffs_if_needed(data["objective_coeffs"], variables)

    objective_sense = str(data["objective_sense"]).lower().strip()
    if objective_sense not in ("max", "min"):
        raise ValueError("objective_sense debe ser 'max' o 'min'")

    constraints = []
    for i, r in enumerate(data["constraints"], 1):
        coeffs = [_as_float(v) for v in r["coeffs"]]
        if len(coeffs) != len(variables):
            raise ValueError(f"La restriccion {i} no tiene {len(variables)} coeficientes")

        sense = str(r["sense"]).strip()
        if sense not in ("<=", ">=", "="):
            raise ValueError(f"Sentido invalido en restriccion {i}: {sense}")

        constraints.append(
            Constraint(
                name=str(r.get("name") or f"R{i}"),
                coeffs=coeffs,
                sense=sense,
                rhs=_as_float(r["rhs"]),
            )
        )

    return LPModel(
        problem_name=str(data.get("problem_name") or "Problema"),
        objective_sense=objective_sense,
        variables=variables,
        objective_coeffs=c,
        constraints=constraints,
    )


def _fmt_num(x: float, digits: int = 4) -> str:
    if abs(x) < 1e-9:
        x = 0.0
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.{digits}f}".rstrip("0").rstrip(".")


def _flip_sense(sense: str) -> str:
    return {"<=": ">=", ">=": "<=", "=": "="}[sense]


def _linear_expression(coeffs: List[float], variables: List[str]) -> str:
    parts = []
    for a, v in zip(coeffs, variables):
        if abs(a) <= 1e-12:
            continue
        sign = "+" if a >= 0 else "-"
        mag = abs(a)
        if abs(mag - 1) < 1e-12:
            term = v
        else:
            term = f"{_fmt_num(mag)}{v}"
        parts.append((sign, term))

    if not parts:
        return "0"

    expr = parts[0][1] if parts[0][0] == "+" else "-" + parts[0][1]
    for sign, term in parts[1:]:
        expr += f" {sign} {term}"
    return expr


def _general_model_lines(lp: LPModel) -> List[str]:
    sentido = "MAX" if lp.objective_sense == "max" else "MIN"
    lines = [
        "Formulacion del modelo general:",
        f"{sentido} Z = {_linear_expression(lp.objective_coeffs, lp.variables)}",
        "Sujeto a:",
    ]

    for i, r in enumerate(lp.constraints, 1):
        lines.append(f"R{i}. {_linear_expression(r.coeffs, lp.variables)} {r.sense} {_fmt_num(r.rhs)}")

    lines.append(", ".join(lp.variables) + " >= 0")
    return lines


def _standard_model_lines(result: Dict[str, Any]) -> List[str]:
    lp = result["lp"]
    std = result["standard"]
    lines = [
        "Formulacion del modelo estandar:",
        f"MAX Z = {_linear_expression(std['c'], std['col_names'])}",
        "Sujeto a:",
    ]

    for i, (row, rhs) in enumerate(zip(std["A"], std["b"]), 1):
        lines.append(f"R{i}. {_linear_expression(row, std['col_names'])} = {_fmt_num(rhs)}")

    lines.append(", ".join(std["col_names"]) + " >= 0")
    if lp.objective_sense == "min":
        lines.append("Nota: el modelo original de minimizacion se transformo a maximizacion para resolverlo con simplex.")
    return lines


def _basic_nonbasic_lines(result: Dict[str, Any]) -> List[str]:
    final_snap = result["iterations"][-1]
    col_names = result["col_names"]
    basic_indexes = final_snap["basis"]
    basic_vars = [col_names[i] for i in basic_indexes]
    nonbasic_vars = [name for j, name in enumerate(col_names) if j not in basic_indexes]

    return [
        "Variables basicas:",
        ", ".join(basic_vars) if basic_vars else "Ninguna",
        "",
        "Variables no basicas:",
        ", ".join(nonbasic_vars) if nonbasic_vars else "Ninguna",
    ]


def _gauss_coeff_text(value: float) -> str:
    if abs(value) < 1e-12:
        return "0"
    if abs(value - 1) < 1e-12:
        return ""
    if abs(value + 1) < 1e-12:
        return "-"
    return _fmt_num(value)


def _gauss_operation_lines(snap: Dict[str, Any]) -> List[str]:
    if snap["status"] != "pivot":
        return ["No se requieren operaciones adicionales."]

    pivot_row = snap["leaving_row"]
    entering = snap["entering"]
    pivot_value = snap["pivot"]
    pivot_coeff = _gauss_coeff_text(1 / pivot_value) or "1"
    lines = [f"F{pivot_row + 1} = ({pivot_coeff})F{pivot_row + 1}"]

    for i, row in enumerate(snap["A"]):
        if i == pivot_row:
            continue
        factor = row[entering]
        if abs(factor) < 1e-12:
            continue
        coeff = _gauss_coeff_text(-factor)
        lines.append(f"F{i + 1} = ({coeff})F{pivot_row + 1} + F{i + 1}")

    return lines


def standardize(lp: LPModel, M: float = 1_000_000.0) -> Dict[str, Any]:
    """
    Convierte el modelo a forma estandar para simplex con metodo de la M grande.
    Para minimizacion se multiplica la funcion objetivo por -1 para resolver como maximizacion.
    """
    n = len(lp.variables)
    obj_for_max = lp.objective_coeffs[:] if lp.objective_sense == "max" else [-v for v in lp.objective_coeffs]

    rows = []
    senses = []
    rhs_values = []

    for r in lp.constraints:
        coeffs = r.coeffs[:]
        rhs = r.rhs
        sense = r.sense

        if rhs < 0:
            coeffs = [-v for v in coeffs]
            rhs = -rhs
            sense = _flip_sense(sense)

        rows.append(coeffs)
        senses.append(sense)
        rhs_values.append(rhs)

    col_names = lp.variables[:]
    c = obj_for_max[:]
    basis = []
    A = [row[:] for row in rows]

    slack_id = 1
    surplus_id = 1
    artificial_id = 1

    for i, sense in enumerate(senses):
        # Variable de holgura para <=
        if sense == "<=":
            for row in A:
                row.append(0.0)
            col_names.append(f"S{slack_id}")
            c.append(0.0)
            A[i][-1] = 1.0
            basis.append(len(col_names) - 1)
            slack_id += 1

        # Variable de excedente y artificial para >=
        elif sense == ">=":
            for row in A:
                row.append(0.0)
            col_names.append(f"E{surplus_id}")
            c.append(0.0)
            A[i][-1] = -1.0
            surplus_id += 1

            for row in A:
                row.append(0.0)
            col_names.append(f"A{artificial_id}")
            c.append(-M)
            A[i][-1] = 1.0
            basis.append(len(col_names) - 1)
            artificial_id += 1

        # Variable artificial para =
        elif sense == "=":
            for row in A:
                row.append(0.0)
            col_names.append(f"A{artificial_id}")
            c.append(-M)
            A[i][-1] = 1.0
            basis.append(len(col_names) - 1)
            artificial_id += 1

    return {
        "A": A,
        "b": rhs_values,
        "c": c,
        "col_names": col_names,
        "basis": basis,
        "M": M,
        "original_n": n,
        "obj_for_max": obj_for_max,
    }


def compute_tableau(A, b, c, basis):
    cb = [c[j] for j in basis]

    zj = []
    for j in range(len(c)):
        zj.append(sum(cb[i] * A[i][j] for i in range(len(A))))

    z_rhs = sum(cb[i] * b[i] for i in range(len(A)))
    cj_zj = [c[j] - zj[j] for j in range(len(c))]

    return cb, zj, z_rhs, cj_zj


def pivot(A, b, basis, pivot_row, pivot_col):
    p = A[pivot_row][pivot_col]
    if abs(p) < 1e-12:
        raise ZeroDivisionError("Pivote cero")

    # Convertir el pivote en 1
    A[pivot_row] = [v / p for v in A[pivot_row]]
    b[pivot_row] /= p

    # Hacer ceros en la columna pivote
    for i in range(len(A)):
        if i == pivot_row:
            continue
        factor = A[i][pivot_col]
        if abs(factor) > 1e-12:
            A[i] = [A[i][j] - factor * A[pivot_row][j] for j in range(len(A[i]))]
            b[i] -= factor * b[pivot_row]

    old = basis[pivot_row]
    basis[pivot_row] = pivot_col
    return old


def simplex(lp: LPModel, max_iter: int = 100, tol: float = 1e-9) -> Dict[str, Any]:
    std = standardize(lp)
    A = [row[:] for row in std["A"]]
    b = std["b"][:]
    c = std["c"][:]
    basis = std["basis"][:]
    col_names = std["col_names"]
    iterations = []

    for it in range(max_iter + 1):
        cb, zj, z_rhs, cj_zj = compute_tableau(A, b, c, basis)

        # Entra la columna con mayor Cj - Zj positivo
        entering = None
        best = tol
        for j, val in enumerate(cj_zj):
            if val > best:
                best = val
                entering = j

        ratios = []
        if entering is not None:
            for i in range(len(A)):
                if A[i][entering] > tol:
                    ratios.append(b[i] / A[i][entering])
                else:
                    ratios.append(math.inf)

        snapshot = {
            "iter": it,
            "A": [row[:] for row in A],
            "b": b[:],
            "basis": basis[:],
            "cb": cb[:],
            "zj": zj[:],
            "z_rhs": z_rhs,
            "cj_zj": cj_zj[:],
            "entering": entering,
            "ratios": ratios[:],
        }

        if entering is None:
            snapshot["status"] = "optimo"
            iterations.append(snapshot)
            break

        if all(math.isinf(r) for r in ratios):
            snapshot["status"] = "no_acotado"
            iterations.append(snapshot)
            raise RuntimeError("Problema no acotado")

        leaving_row = min((r, i) for i, r in enumerate(ratios) if not math.isinf(r))[1]
        snapshot["leaving_row"] = leaving_row
        snapshot["leaving"] = basis[leaving_row]
        snapshot["pivot"] = A[leaving_row][entering]
        snapshot["status"] = "pivot"
        iterations.append(snapshot)

        pivot(A, b, basis, leaving_row, entering)

    else:
        raise RuntimeError("Se alcanzo el maximo de iteraciones")

    solution = {name: 0.0 for name in col_names}
    for i, bi in enumerate(basis):
        solution[col_names[bi]] = b[i]

    artificial_positive = {k: v for k, v in solution.items() if k.startswith("A") and abs(v) > 1e-7}
    if artificial_positive:
        raise RuntimeError(f"El problema parece infactible; artificiales positivas: {artificial_positive}")

    x = {lp.variables[i]: solution.get(lp.variables[i], 0.0) for i in range(len(lp.variables))}
    original_z = sum(lp.objective_coeffs[i] * x[lp.variables[i]] for i in range(len(lp.variables)))

    return {
        "lp": lp,
        "standard": std,
        "iterations": iterations,
        "solution": x,
        "all_solution": solution,
        "objective_value": original_z,
        "col_names": col_names,
        "c": c,
    }


def print_model(lp: LPModel):
    print("\nFORMULACION MODELO GENERAL")
    sentido = "Max." if lp.objective_sense == "max" else "Min."
    terms = _linear_expression(lp.objective_coeffs, lp.variables)
    print(f"F.O : {sentido} Z = {terms}")
    print("C.S.R")

    for i, r in enumerate(lp.constraints, 1):
        lhs = _linear_expression(r.coeffs, lp.variables)
        print(f"{i}) {lhs} {r.sense} {_fmt_num(r.rhs)}   {r.name}")

    print("Xi >= 0")


def print_table(result: Dict[str, Any], snap: Dict[str, Any]):
    col_names = result["col_names"]
    c = result["c"]
    basis = snap["basis"]

    print(f"\nTABLA SIMPLEX - ITERACION {snap['iter']}")
    print("Cj        " + "  ".join(f"{_fmt_num(v):>10}" for v in c) + "       b")
    print("Base Cb   " + "  ".join(f"{name:>10}" for name in col_names) + "       b      razon")

    for i, row in enumerate(snap["A"]):
        ratio = "-" if not snap["ratios"] or math.isinf(snap["ratios"][i]) else _fmt_num(snap["ratios"][i])
        print(
            f"{col_names[basis[i]]:>4} { _fmt_num(snap['cb'][i]):>4} "
            + "  ".join(f"{_fmt_num(v):>10}" for v in row)
            + f"  {_fmt_num(snap['b'][i]):>8}  {ratio:>8}"
        )

    print("Zj        " + "  ".join(f"{_fmt_num(v):>10}" for v in snap["zj"]) + f"  {_fmt_num(snap['z_rhs']):>8}")
    print("Cj-Zj     " + "  ".join(f"{_fmt_num(v):>10}" for v in snap["cj_zj"]))

    if snap["status"] == "pivot":
        print(f"Variable que entra: {col_names[snap['entering']]}")
        print(f"Variable que sale: {col_names[snap['leaving']]}")
        print(f"Pivote: {_fmt_num(snap['pivot'])}")
    else:
        print("Termina: todos los Cj - Zj <= 0. Solucion optima.")


def print_solution(result: Dict[str, Any]):
    print("\nSOLUCION OPTIMA")
    for v, val in result["solution"].items():
        print(f"{v} = {_fmt_num(val)}")
    print(f"Z = {_fmt_num(result['objective_value'])}")


def solve_problem_text(problem_text: str, use_groq: bool = True, json_model: Optional[Dict[str, Any]] = None):
    if use_groq:
        lp = extract_lp_with_groq(problem_text)
    else:
        if json_model is None:
            raise ValueError("Si use_groq=False debes pasar json_model")
        lp = lp_from_dict(json_model)

    print_model(lp)
    result = simplex(lp)

    for snap in result["iterations"]:
        print_table(result, snap)

    print_solution(result)
    return result


def export_steps_to_xlsx(result: Dict[str, Any], path: str = "simplex_resultado.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("Para exportar a Excel instala: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Simplex"

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="EAF4EA")
    pivot_fill = PatternFill("solid", fgColor="FFF59D")
    thin = Side(style="thin", color="C9C9C9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    lp = result["lp"]

    ws.cell(row, 1, lp.problem_name).font = Font(bold=True, size=14)
    row += 2

    ws.cell(row, 1, "Formulacion Modelo General").font = Font(bold=True)
    ws.cell(row, 1).fill = title_fill
    row += 1

    ws.cell(row, 1, "Funcion objetivo")
    ws.cell(row, 2, f"{lp.objective_sense.upper()} Z = {_linear_expression(lp.objective_coeffs, lp.variables)}")
    row += 1

    ws.cell(row, 1, "Restricciones")
    row += 1

    for i, r in enumerate(lp.constraints, 1):
        ws.cell(row, 1, f"R{i}")
        ws.cell(row, 2, f"{_linear_expression(r.coeffs, lp.variables)} {r.sense} {_fmt_num(r.rhs)}")
        ws.cell(row, 3, r.name)
        row += 1

    row += 1

    ws.cell(row, 1, "Formulacion Modelo Estandar").font = Font(bold=True)
    ws.cell(row, 1).fill = title_fill
    row += 1

    std = result["standard"]
    ws.cell(row, 1, "Funcion objetivo")
    ws.cell(row, 2, f"MAX Z = {_linear_expression(std['c'], std['col_names'])}")
    row += 1

    ws.cell(row, 1, "Restricciones")
    row += 1

    for i, (coeffs, rhs) in enumerate(zip(std["A"], std["b"]), 1):
        ws.cell(row, 1, f"R{i}")
        ws.cell(row, 2, f"{_linear_expression(coeffs, std['col_names'])} = {_fmt_num(rhs)}")
        row += 1

    ws.cell(row, 1, "No negatividad")
    ws.cell(row, 2, ", ".join(std["col_names"]) + " >= 0")
    row += 2

    for snap in result["iterations"]:
        ws.cell(row, 1, f"Tabla simplex - Iteracion {snap['iter']}").font = Font(bold=True)
        ws.cell(row, 1).fill = title_fill
        row += 1

        headers = ["Var. Basica", "Cb"] + result["col_names"] + ["b", "Razon"]
        ops_col = len(headers) + 3
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row, col, h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        ops_header = ws.cell(row, ops_col, "Procedimiento Gauss")
        ops_header.font = Font(bold=True)
        ops_header.fill = header_fill
        ops_header.border = border
        row += 1

        table_start_row = row
        for i, vec in enumerate(snap["A"]):
            vals = [result["col_names"][snap["basis"][i]], snap["cb"][i]] + vec + [snap["b"][i]]
            if snap["ratios"]:
                vals.append("-" if math.isinf(snap["ratios"][i]) else snap["ratios"][i])
            else:
                vals.append("-")

            for col, val in enumerate(vals, 1):
                ws.cell(row, col, val)
                ws.cell(row, col).border = border
            row += 1

        for offset, op in enumerate(_gauss_operation_lines(snap)):
            op_cell = ws.cell(table_start_row + offset, ops_col, op)
            op_cell.border = border

        if snap["status"] == "pivot":
            pivot_row_excel = table_start_row + snap["leaving_row"]
            pivot_col_excel = 3 + snap["entering"]
            pivot_cell = ws.cell(pivot_row_excel, pivot_col_excel)
            pivot_cell.fill = pivot_fill

        # Fila Zj
        ws.cell(row, 1, "Zj").font = Font(bold=True)
        zj_vals = ["", ""] + snap["zj"] + [snap["z_rhs"], ""]
        for col, val in enumerate(zj_vals, 1):
            ws.cell(row, col, val)
            ws.cell(row, col).border = border
        row += 1

        # Fila Cj - Zj
        ws.cell(row, 1, "Cj - Zj").font = Font(bold=True)
        cjzj_vals = ["", ""] + snap["cj_zj"] + ["", ""]
        for col, val in enumerate(cjzj_vals, 1):
            ws.cell(row, col, val)
            ws.cell(row, col).border = border
        row += 1

        if snap["status"] == "pivot":
            ws.cell(row, 1, "Variable que entra")
            ws.cell(row, 2, result["col_names"][snap["entering"]])
            ws.cell(row, 3, "Variable que sale")
            ws.cell(row, 4, result["col_names"][snap["leaving"]])
            ws.cell(row, 5, "Pivote")
            ws.cell(row, 6, snap["pivot"])
            row += 2
        else:
            ws.cell(row, 1, "Solucion optima: todos los Cj - Zj <= 0")
            row += 2

    final_basis = result["iterations"][-1]["basis"]
    basic_vars = [result["col_names"][i] for i in final_basis]
    nonbasic_vars = [name for j, name in enumerate(result["col_names"]) if j not in final_basis]

    ws.cell(row, 1, "Variables basicas").font = Font(bold=True)
    ws.cell(row, 1).fill = title_fill
    ws.cell(row, 2, ", ".join(basic_vars) if basic_vars else "Ninguna")
    row += 1

    ws.cell(row, 1, "Variables no basicas").font = Font(bold=True)
    ws.cell(row, 1).fill = title_fill
    ws.cell(row, 2, ", ".join(nonbasic_vars) if nonbasic_vars else "Ninguna")
    row += 2

    ws.cell(row, 1, "Solucion optima").font = Font(bold=True, size=12)
    ws.cell(row, 1).fill = title_fill
    row += 1

    for v, val in result["solution"].items():
        ws.cell(row, 1, v)
        ws.cell(row, 2, val)
        row += 1

    ws.cell(row, 1, "Z")
    ws.cell(row, 2, result["objective_value"])

    for col in range(1, ws.max_column + 1):
        letter = ws.cell(1, col).column_letter
        ws.column_dimensions[letter].width = 16

    wb.save(path)
    return path


def _format_result_summary(result: Dict[str, Any], excel_path: Optional[str] = None) -> str:
    lp = result["lp"]
    lines = [
        f"Problema: {lp.problem_name}",
        "",
    ]

    lines.extend(_general_model_lines(lp))
    lines.extend([""])
    lines.extend(_standard_model_lines(result))
    lines.extend([""])
    lines.extend(_basic_nonbasic_lines(result))
    lines.extend(["", "Solucion optima:"])
    for v, val in result["solution"].items():
        lines.append(f"{v} = {_fmt_num(val)}")
    lines.append(f"Z = {_fmt_num(result['objective_value'])}")
    lines.append(f"Iteraciones simplex: {len(result['iterations'])}")

    if excel_path:
        lines.extend(["", f"Excel generado: {excel_path}"])

    return "\n".join(lines)


def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk

    root = tk.Tk()
    root.title("Solver Simplex con Groq")
    root.geometry("980x720")
    root.minsize(820, 600)

    result_state = {"result": None, "excel_path": None}

    root.columnconfigure(0, weight=1)
    root.rowconfigure(2, weight=1)
    root.rowconfigure(4, weight=1)

    header = ttk.Frame(root, padding=(14, 12, 14, 6))
    header.grid(row=0, column=0, sticky="ew")
    header.columnconfigure(0, weight=1)

    ttk.Label(header, text="Metodo Simplex", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w")
    ttk.Label(
        header,
        text="Pega el enunciado del problema, resuelvelo y guarda el Excel con las tablas.",
    ).grid(row=1, column=0, sticky="w", pady=(3, 0))

    config_frame = ttk.LabelFrame(root, text="Configuracion de Groq", padding=10)
    config_frame.grid(row=1, column=0, sticky="ew", padx=14, pady=(8, 4))
    config_frame.columnconfigure(1, weight=1)

    api_key_var = tk.StringVar(value=get_groq_api_key())
    api_key_status_var = tk.StringVar()
    show_api_key_var = tk.BooleanVar(value=False)

    ttk.Label(config_frame, text="API key").grid(row=0, column=0, sticky="w", padx=(0, 8))
    api_key_entry = ttk.Entry(config_frame, textvariable=api_key_var, show="*", width=60)
    api_key_entry.grid(row=0, column=1, sticky="ew")

    def toggle_api_key_visibility():
        api_key_entry.configure(show="" if show_api_key_var.get() else "*")

    ttk.Checkbutton(
        config_frame,
        text="Mostrar",
        variable=show_api_key_var,
        command=toggle_api_key_visibility,
    ).grid(row=0, column=2, padx=8)

    def persist_api_key(show_success: bool = True) -> bool:
        api_key = api_key_var.get().strip()
        if not api_key:
            api_key_status_var.set("Ingresa una API key valida.")
            if show_success:
                messagebox.showwarning("Falta la API key", "Escribe la API key de Groq antes de guardarla.")
            return False

        try:
            set_groq_api_key(api_key)
        except Exception as exc:
            api_key_status_var.set("No se pudo guardar la API key.")
            if show_success:
                messagebox.showerror("Error al guardar", str(exc))
            return False

        api_key_status_var.set(f"API key guardada en {APP_CONFIG_PATH}")
        if show_success:
            messagebox.showinfo("Configuracion guardada", "La API key de Groq quedo guardada en este equipo.")
        return True

    ttk.Button(config_frame, text="Guardar API key", command=persist_api_key).grid(row=0, column=3, padx=(8, 0))
    ttk.Label(
        config_frame,
        text="Si existe GROQ_API_KEY en el sistema, esa prioridad se mantiene sobre la guardada aqui.",
    ).grid(row=1, column=0, columnspan=4, sticky="w", pady=(8, 0))
    ttk.Label(config_frame, textvariable=api_key_status_var).grid(row=2, column=0, columnspan=4, sticky="w", pady=(4, 0))

    api_key_status_var.set(
        f"API key cargada desde {APP_CONFIG_PATH}" if api_key_var.get().strip() else "Aun no hay una API key guardada."
    )

    input_frame = ttk.LabelFrame(root, text="Enunciado del problema", padding=10)
    input_frame.grid(row=2, column=0, sticky="nsew", padx=14, pady=8)
    input_frame.columnconfigure(0, weight=1)
    input_frame.rowconfigure(0, weight=1)

    problem_text = scrolledtext.ScrolledText(input_frame, wrap=tk.WORD, height=12, font=("Segoe UI", 10))
    problem_text.grid(row=0, column=0, sticky="nsew")

    actions = ttk.Frame(root, padding=(14, 4, 14, 4))
    actions.grid(row=3, column=0, sticky="ew")
    actions.columnconfigure(2, weight=1)

    status_var = tk.StringVar(value="Listo")
    solve_button = ttk.Button(actions, text="Resolver y generar Excel")
    save_button = ttk.Button(actions, text="Descargar Excel", state=tk.DISABLED)
    clear_button = ttk.Button(actions, text="Limpiar")

    solve_button.grid(row=0, column=0, padx=(0, 8))
    save_button.grid(row=0, column=1, padx=(0, 8))
    clear_button.grid(row=0, column=2, sticky="w")
    ttk.Label(actions, textvariable=status_var).grid(row=0, column=3, sticky="e")

    output_frame = ttk.LabelFrame(root, text="Resultado", padding=10)
    output_frame.grid(row=4, column=0, sticky="nsew", padx=14, pady=(8, 14))
    output_frame.columnconfigure(0, weight=1)
    output_frame.rowconfigure(0, weight=1)

    output_text = scrolledtext.ScrolledText(output_frame, wrap=tk.WORD, height=10, font=("Consolas", 10), state=tk.DISABLED)
    output_text.grid(row=0, column=0, sticky="nsew")

    def set_output(text: str):
        output_text.configure(state=tk.NORMAL)
        output_text.delete("1.0", tk.END)
        output_text.insert(tk.END, text)
        output_text.configure(state=tk.DISABLED)

    def set_busy(is_busy: bool):
        solve_button.configure(state=tk.DISABLED if is_busy else tk.NORMAL)
        clear_button.configure(state=tk.DISABLED if is_busy else tk.NORMAL)
        if is_busy:
            save_button.configure(state=tk.DISABLED)

    def solve_worker(text: str):
        try:
            result = solve_problem_text(text, use_groq=True)
            excel_path = export_steps_to_xlsx(result, "simplex_resultado.xlsx")
        except Exception as exc:
            root.after(0, lambda: finish_error(exc))
            return

        root.after(0, lambda: finish_success(result, excel_path))

    def finish_success(result: Dict[str, Any], excel_path: str):
        result_state["result"] = result
        result_state["excel_path"] = excel_path
        set_output(_format_result_summary(result, excel_path))
        save_button.configure(state=tk.NORMAL)
        status_var.set("Resuelto correctamente")
        set_busy(False)
        messagebox.showinfo("Simplex", "Problema resuelto. Ya puedes descargar el Excel.")

    def finish_error(exc: Exception):
        result_state["result"] = None
        result_state["excel_path"] = None
        set_output(f"Error:\n{exc}")
        status_var.set("Error al resolver")
        set_busy(False)
        messagebox.showerror("No se pudo resolver", str(exc))

    def solve_from_input():
        text = problem_text.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("Falta el enunciado", "Escribe o pega el problema antes de resolver.")
            return

        if not get_groq_api_key():
            if not persist_api_key(show_success=False):
                messagebox.showwarning("Falta la API key", "Configura la API key de Groq antes de resolver.")
                return

        set_busy(True)
        status_var.set("Resolviendo con Groq...")
        set_output("Procesando el enunciado y generando el archivo Excel...")
        threading.Thread(target=solve_worker, args=(text,), daemon=True).start()

    def save_excel():
        result = result_state["result"]
        if result is None:
            messagebox.showwarning("Sin resultado", "Primero resuelve un problema.")
            return

        path = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            initialfile="simplex_resultado.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        try:
            export_steps_to_xlsx(result, path)
        except Exception as exc:
            messagebox.showerror("No se pudo guardar", str(exc))
            return

        result_state["excel_path"] = path
        set_output(_format_result_summary(result, path))
        status_var.set(f"Excel guardado: {path}")
        messagebox.showinfo("Excel guardado", f"Archivo guardado en:\n{path}")

    def clear_all():
        problem_text.delete("1.0", tk.END)
        result_state["result"] = None
        result_state["excel_path"] = None
        save_button.configure(state=tk.DISABLED)
        status_var.set("Listo")
        set_output("")

    solve_button.configure(command=solve_from_input)
    save_button.configure(command=save_excel)
    clear_button.configure(command=clear_all)

    root.mainloop()


if __name__ == "__main__":
    run_gui()
